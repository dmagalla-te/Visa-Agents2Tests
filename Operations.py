import os
import json
from Connector import get_data, post_data
from openpyxl import load_workbook
from datetime import datetime

def timestamp():

    date_time_now = datetime.now()
    return date_time_now.strftime("%d%b%y-%H%M%S")

def set_logs(file_path, logs):
    
    logs = str(logs)
    file_name = file_path + '.log'
    f = open(file_name, 'a+')  # open file in append mode
    f.write(logs)
    f.close()

    return "Logs saved"


def provision_agents(te_tests, OAUTH):

    headers = {
        'Authorization' : 'Bearer ' + OAUTH,
        'Content-Type' : 'application/json'
    }

    for i in te_tests.values():

        url = "https://api.thousandeyes.com/v6/tests/http-server/%s/update.json?aid=%s" % (i[1], i[0]) 

        agents = i[2]

        new_agents = []

        print(agents)

        if agents:

            for agent in agents:

                new_agents.append({"agentId": agent})

            payload = {"agents": new_agents, "enabled": 1} #asgin agents and enable test
            print(payload)

            print(post_data(headers, endp_url=url, payload=json.dumps(payload)))
        
        #if the test does not have any agents assigned to it
        else:
            
            payload = {"agents": [], "enabled": 0} #just disable the test
            print(payload)

            print(post_data(headers, endp_url=url, payload=json.dumps(payload)))


######################
########### account, testname , source
###########
def get_tests(headers, aid):

    tests_endp = "https://api.thousandeyes.com/v6/tests.json"

    tests = get_data(headers, tests_endp, params={"aid":aid})

    if 'test' in tests:
        
        tests_relation = {}

        for test in tests['test']:

            tests_relation.update({test.get('testName'): test.get("testId")})

        return tests_relation
    
    else:

        return False



def read_excel():

    tests = {}
    workbook = load_workbook(filename='CIT.xlsx')
    sheet = workbook.active


    for row in sheet.iter_rows(min_row=2, max_col=5, max_row=sheet.max_row, values_only=True):   

        if row[0] is None: #si no hay email no hacemos nada 
            continue  
        
        account_group = row[0]
        test_id = row[1]
        agent_name = row[2]

        if account_group in tests: 

            if test_id in tests[account_group]:
                
                tests[account_group][test_id].append(agent_name)

            else:

                tests[account_group][test_id] = [agent_name]
        else:
            tests[account_group] = {test_id:[agent_name]}


    return tests



def get_agents(headers, aid):

    endp_url2 = "https://api.thousandeyes.com/v6/agents.json"

    agents = get_data(headers, endp_url2, params={"aid":aid})

    if 'agents' in agents:
        
        agent_relation = {}

        for agent in agents['agents']:

            agent_relation.update({agent.get('agentName'): agent.get("agentId")})

        return agent_relation
    
    else:

        return False
    


def agent_ids(OAUTH, test_agents):

    headers = {
        'Authorization' : 'Bearer ' + OAUTH,
        'Content-Type' : 'application/json'
    }

    logs = ''


    for aid, inner_dict in test_agents.items():
        print('ACCOUNT GROUP:', aid)
        agents = get_agents(headers, aid) #dict relacion AGENT_NAME: AGENT_ID
        tests = get_tests(headers, aid) #dict relacion TEST_NAME: TEST_ID

        for test_name, agents_names in inner_dict.items():
            print('TEST', test_name)
            agents_ids = []

            for name in agents_names:
                #hacer lista de agent ids
                my_variable = agents.get(name)

                if isinstance(my_variable, int):
                    agents_ids.append({"agentId": my_variable})


            print('AGENTS', agents_ids)

            test_id = tests.get(test_name)
            #get test details para sacar test type
            test_details_endp = "https://api.thousandeyes.com/v6/tests/%s.json" % test_id
            test_details = get_data(headers, test_details_endp, params={"aid":aid})

            test_type = test_details['test'][0]['type']


            #Una vez que tengamos la lista de los agent ids post a los test
            url = "https://api.thousandeyes.com/v6/tests/%s/%s/update.json?aid=%s" % (test_type,test_id, aid)
            print(url)
            payload = {"agents": agents_ids, "enabled": 1} #asgin agents and enable test
            
            if agents_ids != []:
                try:
                    status_code = post_data(headers, endp_url=url, payload=json.dumps(payload))
                    logs += "\n" + timestamp() + "-> TEST: " + test_name +' Status code: ' + str(status_code) + ' Url: ' + url


                except:
                    logs += "\n" + timestamp() + "-> This test could not be updated: " + test_name + ' Url: ' + url
    
    
    set_logs(file_path="Cardinal", logs=logs)

            

